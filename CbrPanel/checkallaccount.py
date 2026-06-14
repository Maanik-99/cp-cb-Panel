import json
import requests
import urllib3
from pathlib import Path
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE_DIR = Path(__file__).parent

CPANEL_WORKBOOK_FILE = BASE_DIR / "CPanel.xlsx"
OUTPUT_FILE = BASE_DIR / "fetched_emails.txt"


def login_to_panel(base_url, username, password):
    session = requests.Session()
    session.verify = False

    try:
        if base_url.endswith("/"):
            base_url = base_url[:-1]

        login_page = f"{base_url}/login"

        session.get(login_page, timeout=20)

        csrf = session.cookies.get("csrftoken", "")

        headers = {
            "X-CSRFToken": csrf,
            "Referer": login_page,
            "Origin": base_url,
            "User-Agent": "Mozilla/5.0",
            "X-Requested-With": "XMLHttpRequest",
            "Content-Type": "application/json"
        }

        payload = {
            "username": username,
            "password": password,
            "languageSelection": "english"
        }

        response = session.post(
            f"{base_url}/verifyLogin",
            headers=headers,
            json=payload,
            timeout=20
        )

        text = response.text.lower()

        if '"loginstatus": 1' in text or '"loginstatus":1' in text:
            print(f"[+] Login Success: {base_url}")
            return session

        print(f"[-] Login Failed: {base_url}")
        print(response.text[:500])

        return None

    except Exception as e:
        print(f"[-] Login Error: {e}")
        return None


def fetch_all_emails(session, base_url, domain):

    try:
        csrf = session.cookies.get("csrftoken", "")

        headers = {
            "X-CSRFToken": csrf,
            "Referer": f"{base_url}/email/listEmails",
            "Origin": base_url,
            "User-Agent": "Mozilla/5.0",
            "X-Requested-With": "XMLHttpRequest",
            "Content-Type": "application/json"
        }

        payload = {
            "selectedDomain": domain
        }

        response = session.post(
            f"{base_url}/email/fetchEmails",
            headers=headers,
            json=payload,
            timeout=30
        )

        result = response.json()

        if result.get("status") != 1:
            print(f"[!] Fetch Failed ({domain})")
            print(result)
            return []

        raw_data = result.get("data", "[]")

        if isinstance(raw_data, str):
            try:
                raw_data = json.loads(raw_data)
            except:
                raw_data = []

        emails = []

        for item in raw_data:

            email = item.get("email", "")
            usage = item.get("DiskUsage", "")

            if email:
                emails.append({
                    "email": email,
                    "DiskUsage": usage
                })

        return emails

    except Exception as e:
        print(f"Fetch Error ({domain}): {e}")
        return []


def main():

    if not CPANEL_WORKBOOK_FILE.exists():
        print("CPanel.xlsx not found")
        return

    wb = load_workbook(CPANEL_WORKBOOK_FILE)
    ws = wb.active

    headers = [
        str(cell.value).strip().upper()
        if cell.value else ""
        for cell in ws[1]
    ]

    try:
        idx_base_url = headers.index("BASE_URL")
        idx_domain = headers.index("EMAIL_DOMAIN")
        idx_username = headers.index("LOGIN_USERNAME")
        idx_password = headers.index("LOGIN_PASSWORD")
    except ValueError as e:
        print("Missing Column:", e)
        print("Detected Headers:", headers)
        return

    all_accounts = []

    last_base_url = None
    last_domain = None

    for row in ws.iter_rows(min_row=2, values_only=True):

        base_url = row[idx_base_url]
        domain = row[idx_domain]
        username = row[idx_username]
        password = row[idx_password]

        if not base_url and last_base_url:
            base_url = last_base_url

        if not domain and last_domain:
            domain = last_domain

        if base_url:
            last_base_url = base_url

        if domain:
            last_domain = domain

        if not base_url or not domain:
            continue

        print(f"\nProcessing {base_url}")
        print(f"Domain: {domain}")

        session = login_to_panel(
            str(base_url).strip(),
            str(username).strip(),
            str(password).strip()
        )

        if not session:
            continue

        accounts = fetch_all_emails(
            session,
            str(base_url).strip(),
            str(domain).strip()
        )

        print(f"FOUND: {len(accounts)}")

        for account in accounts:

            account["server"] = str(base_url).strip()
            account["domain"] = str(domain).strip()

            all_accounts.append(account)

            print(
                f"{account['email']} | "
                f"{account['DiskUsage']}"
            )

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:

        for account in all_accounts:

            f.write(
                f"{account['email']} | "
                f"{account['DiskUsage']} | "
                f"{account['domain']} | "
                f"{account['server']}\n"
            )

    print("\n====================================")
    print(f"Total Emails Found: {len(all_accounts)}")
    print(f"Saved To: {OUTPUT_FILE}")
    print("====================================")


if __name__ == "__main__":
    main()